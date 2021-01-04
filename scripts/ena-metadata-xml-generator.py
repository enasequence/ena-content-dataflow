from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import lxml.etree
import lxml.builder
import xlrd
from yattag import Doc, indent
import argparse, hashlib, os, subprocess, sys, time
from datetime import datetime

#wb = load_workbook ('uploader_tool_metadata_v3_raw_reads_test2.xlsx')
#sh = wb.active
wb = load_workbook("uploader_tool_metadata_v3_raw_reads_test3.xlsx")
ws = wb.worksheets[0]

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'


doc.asis(xml_header)
#doc.asis(xml_schema)
#tags = [n.replace(" ", "").lower() for n in sh.row_values(0)]
#iterate through rows 

#print (ws[2])
    # Use ws.max_row for all rows
i=5
for row in ws.iter_rows(min_row=6, min_col=12, max_col=41, values_only=True):
    found = False
    for x in row:
        if x != None:
            found = True
    if found == True:
        #print(row)
        first = row[0:5]
        all = row[5:]
        #print(first)
        #print(all)
        with tag('SAMPLE_SET'):
            with tag('SAMPLE', alias=first[0]):
                with tag("TITLE"):
                    text(first[3])
                with tag('SAMPLE_NAME'):
                    with tag("TAXON_ID"):
                        text(first[1])
                    with tag("SCIENTIFIC_NAME"):
                        text(first[2])
                with tag("DESCRIPTION"):
                    text(first[4])
                    with tag('SAMPLE_ATTRIBUTES'):
                        for x in all:
                            if x != None:
                                with tag("SAMPLE_ATTRIBUTE"):
                                    with tag("TAG"):
                                        text(ws[2][row.index(x)+11].value)
                                    with tag("VALUE"):
                                        text(str(x))
                        with tag("SAMPLE_ATTRIBUTE"):
                            with tag("TAG"):
                                text("ENA-CHECKLIST")
                            with tag("VALUE"):
                                text('ERC000033')

                    i=i+1


result = indent(
    doc.getvalue(),
    indentation = '    ',
    indent_text = True
)

with open("sample.xml", "w") as f:
    f.write(result)

#creating the submission xml

# Create Yattag doc, tag and text objects
doc, tag, text = Doc().tagtext()
xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
#xml_schema = '<xs:schema xmlns:xs="http://www.w3.org/2001/XMLSchema"></xs:schema>'

doc.asis(xml_header)
#doc.asis(xml_schema)

with tag('SUBMISSION_SET'):
    with tag('SUBMISSION'):
        with tag("ACTIONS"):
            with tag('ACTION'):
                doc.stag('ADD')

result_s = indent(
    doc.getvalue(),
    indentation='    ',
    indent_text=True
)

with open("submission.xml", "w") as f:
    f.write(result_s)
######################################
username = 'Webin-56352'
password = 'Dp9HOUhhap'
action = 'ADD'
#######################
# submission command
command = 'curl -u {}:{} -F "SUBMISSION=@submission.xml" -F "SAMPLE=@sample.xml" "https://wwwdev.ebi.ac.uk/ena/submit/drop-box/submit/"'.format(username, password)
print (command)
sp = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
out, err = sp.communicate()